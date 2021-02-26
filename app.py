from flask import Flask, render_template, redirect
from forms import SubmitForm
import xlsxwriter
import time 
import datetime as dt
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Color, PatternFill, Alignment
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import RichTextProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.layout import Layout, ManualLayout
import yaml
import requests

application = Flask(__name__)
application.config['SECRET_KEY'] = 'you-will-never-guess'

@application.route("/", methods=['GET', 'POST'])
def enterData():
    form = SubmitForm()
    if form.validate_on_submit():
        parseData(form.inputData.data)
        return redirect('/')
    return render_template('submit.html', form=form)

def parseData(inputData):
    myData = []
    for x in inputData.split("-"):
        myData.append(x)
    writeExcel(myData)

def getConfig():
    with open("/home/ec2-user/CM-APP/config.yml", "r") as ymlfile:
        cfg = yaml.load(ymlfile)
    return cfg

def saveRefreshToken(token):
    cfg = getConfig()
    cfg["refreshToken"] = token
    with open("/home/ec2-user/CM-APP/config.yml", "w") as ymlfile:
        yaml.dump(cfg, ymlfile)

class StrictDict(dict):
    def __setitem__(self, key, value):
        if key in self:
            dict.__setitem__(self, key, value)


def writeExcel(myData):
    initialDateString = myData.pop(0)
    initialDate = date(int(initialDateString.split(",")[0]), int(initialDateString.split(",")[1]), int(initialDateString.split(",")[2]))
    
    workbook = xlsxwriter.Workbook('MC_Report.xlsx')
    worksheet = workbook.add_worksheet(str(initialDate))
    headers = ['Day', 'Date', 'T (ºC)', 'CM', 'M-Quantity', 'Weight']
    row = 0
    column = 0

    for header in headers:
        worksheet.write(row, column, header)
        column += 1

    row = 1
    column = 2
    counter = 1
    incrementalDays = 0
    cellFormat = workbook.add_format()
    cellFormatNumber = workbook.add_format()
    cellFormat.set_num_format('dd/mm/yy')
    cellFormatNumber.set_num_format('##,##')
    datesRange = {}
    for line in myData:
        worksheet.write(row, 0, counter)
        worksheet.write(row, 1, initialDate+relativedelta(days=+incrementalDays), cellFormat)
        datesRange[str(initialDate+relativedelta(days=+incrementalDays))] = ''
        first_time = True
        for item in line.split(" "):
            if item is not None and item != " " and item != "":
                if first_time:
                    if item != "NA":
                        worksheet.write_number(row, column, float(item.replace(',','.')))
                    column += 1
                    first_time = False
                else:
                    worksheet.write(row, column, item)
                    column += 1
        column = 2
        row += 1
        counter += 1
        incrementalDays += 1

    # To add weight on the Excel.
    stricDateRange = StrictDict(datesRange)
    weightDates = getWeight(stricDateRange)
    row = 1
    column = 5
    for weight in weightDates.values():
        worksheet.write(row, column, weight)
        row += 1

    workbook.close()
    # Get mean weight.
    meanWeight(initialDate,counter-1)

    # Create chart.
    createChart(initialDate, counter-1)

def getWeight(datesRange):
    # Get Config
    cfg = getConfig()

    # Request to refresh token. 
    url = "https://wbsapi.withings.net/v2/oauth2"
    payload='action=requesttoken&client_id='+cfg['clientID']+'&client_secret='+cfg['clientSecret']+'&grant_type=refresh_token&refresh_token='+cfg['refreshToken']
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    response = requests.request("POST", url, headers=headers, data=payload).json()

    # Save new refresh token
    newRefreshToken = response['body']['refresh_token']
    print(newRefreshToken)
    saveRefreshToken(newRefreshToken)

    # Request to get weight.
    headers = {'Authorization': 'Bearer ' + response['body']['access_token']}
    startDate = time.mktime(dt.datetime.strptime(list(datesRange.keys())[0], "%Y-%m-%d").timetuple())
    payload = {'action': 'getmeas', 'maestype': 1, 'category': 1, 'startdate': startDate}

    # List of weights.
    r_getWeight = requests.get('https://wbsapi.withings.net/measure', headers=headers, params=payload).json()
    
    for weight in r_getWeight['body']['measuregrps']:
        datesRange[str(datetime.utcfromtimestamp(weight['date']).strftime('%Y-%m-%d'))] = weight['measures'][0]['value']*0.001
    return datesRange

def meanWeight(initialDate,totalDays):
    wb = load_workbook('MC_Report.xlsx')
    ws = wb[str(initialDate)]

    areThereM = False
    for x in range(totalDays+1):
        if ws.cell(x+2-1,4).value == "M":
            areThereM = True
        if ws.cell(x+2-1,4).value == "D" and areThereM:
            firstValue = x+2
            lastValue = x+8
            break

    weights = []
    for x in range(firstValue, lastValue+1):
        weights.append(ws.cell(x,6).value)
    ws.merge_cells(start_row=firstValue, start_column=7, end_row=lastValue, end_column=7)
    ws.cell(firstValue,7).value = sum(weights)/len(weights)
    ws.cell(firstValue,7).alignment = Alignment(vertical='center')
    wb.save("MC_Report.xlsx")

def createChart(initialDate, totalDays):
    wb = load_workbook('MC_Report.xlsx')
    ws = wb[str(initialDate)]

    chart = ScatterChart()
    # Create label styling.
    axis = CharacterProperties(sz=800)
    rot = openpyxl.drawing.text.RichTextProperties(rot=-5400000)

    # Set axis label styles.
    chart.x_axis.title = 'Date'
    manualLayout = ManualLayout(xMode="factor", yMode="factor", x=-0.47, y=-0.025)
    chart.x_axis.title.layout = Layout(manualLayout = manualLayout)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])
    chart.y_axis.title = 'Temperature'
    
    redMarker = ColorChoice(prstClr='red')
    orangeMarker = ColorChoice(prstClr='orange')
    yellowMarker = ColorChoice(prstClr='yellow')
    lightPinkMarker = ColorChoice(prstClr='lightPink')
    hotPinkMarker = ColorChoice(prstClr='hotPink')
    blackMarker = ColorChoice(prstClr='black')
    blackLine = ColorChoice(prstClr='black')
    lineMarker = LineProperties(solidFill=blackLine)
    
    # To initiate the values of the first serie.
    firstValueSerie = ws.cell(2,4).value
    firstPositionSerie = 2
    
    for x in range(totalDays+1):
        if ws.cell(x+2,4).value == firstValueSerie:
            pass
        else:
            xvalues = Reference(ws, min_col=2, min_row=firstPositionSerie,max_row=str(x+2))
            yvalues = Reference(ws, min_col=3, min_row=firstPositionSerie,max_row=str(x+2))
            if ws.cell(x+2-1,4).value == "M":
                markerProperty = GraphicalProperties(solidFill=redMarker, ln=lineMarker)
            if ws.cell(x+2-1,4).value == "m":
                markerProperty = GraphicalProperties(solidFill=orangeMarker, ln=lineMarker)
            if ws.cell(x+2-1,4).value == "D":
                markerProperty = GraphicalProperties(solidFill=yellowMarker, ln=lineMarker)
            if ws.cell(x+2-1,4).value == "P":
                markerProperty = GraphicalProperties(solidFill=hotPinkMarker, ln=lineMarker)
            if ws.cell(x+2-1,4).value == "p":
                markerProperty = GraphicalProperties(solidFill=lightPinkMarker, ln=lineMarker)
            if ws.cell(x+2-1,4).value == "NA":
                markerProperty = GraphicalProperties(solidFill=blackMarker, ln=lineMarker)

            series = Series(yvalues, xvalues, title_from_data=False)
            series.marker = openpyxl.chart.marker.Marker('circle', spPr=markerProperty)
            series.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill = blackLine)
            chart.series.append(series)
            firstValueSerie = ws.cell(x+2,4).value
            firstPositionSerie = x+2
    
    # Last serie.
    xvalues = Reference(ws, min_col=2, min_row=firstPositionSerie,max_row=totalDays+1)
    yvalues = Reference(ws, min_col=3, min_row=firstPositionSerie,max_row=totalDays+1)
    if ws.cell(x+2-1,4).value == "M":
        markerProperty = GraphicalProperties(solidFill=redMarker, ln=lineMarker)
    if ws.cell(x+2-1,4).value == "m":
        markerProperty = GraphicalProperties(solidFill=orangeMarker, ln=lineMarker)
    if ws.cell(x+2-1,4).value == "D":
        markerProperty = GraphicalProperties(solidFill=yellowMarker, ln=lineMarker)
    if ws.cell(x+2-1,4).value == "P":
        markerProperty = GraphicalProperties(solidFill=hotPinkMarker, ln=lineMarker)
    if ws.cell(x+2-1,4).value == "p":
        markerProperty = GraphicalProperties(solidFill=lightPinkMarker, ln=lineMarker)
    if ws.cell(x+2-1,4).value == "NA":
        markerProperty = GraphicalProperties(solidFill=blackMarker, ln=lineMarker)
    series = Series(yvalues, xvalues, title_from_data=False)
    series.marker = openpyxl.chart.marker.Marker('circle', spPr=markerProperty)
    series.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill = blackLine)
    chart.series.append(series)

    # Remove legend.
    chart.legend= None
    chart.x_axis.majorUnit = 1
    ws.add_chart(chart, 'I7')


    # Fill MC cells according to a legend.
    myRed = openpyxl.styles.colors.Color(rgb="FF0000")
    myYellow = openpyxl.styles.colors.Color(rgb="FFFF33")
    myOrange = openpyxl.styles.colors.Color(rgb="FFA500")
    myLightPink = openpyxl.styles.colors.Color(rgb="FFB6C1")
    myHotPink = openpyxl.styles.colors.Color(rgb="FF69B4")
    myRedFill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=myRed)
    myYellowFill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=myYellow)
    myOrangeFill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=myOrange)
    myLightPinkFill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=myLightPink)
    myHotPinkFill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=myHotPink)

    for x in range(totalDays):
         if ws.cell(x+2,4).value == 'M':
            ws.cell(x+2,4).fill = myRedFill
         if ws.cell(x+2,4).value == 'D':
            ws.cell(x+2,4).fill = myYellowFill
         if ws.cell(x+2,4).value == 'm':
            ws.cell(x+2,4).fill = myOrangeFill
         if ws.cell(x+2,4).value == 'p':
            ws.cell(x+2,4).fill = myLightPinkFill
         if ws.cell(x+2,4).value == 'P':
            ws.cell(x+2,4).fill = myHotPinkFill    

    wb.save("MC_Report.xlsx")
    sendEmail()

def sendEmail():
    
    # Get Config
    cfg = getConfig()

    subject = "MC Report"
    body = "MC Report"
    senderEmail = cfg['senderEmail']
    receiverEmail = cfg['receiverEmail']
    password = cfg['password']

    message = MIMEMultipart()
    message["From"] = senderEmail
    message["To"] = receiverEmail
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    filename = "MC_Report.xlsx"  # In same directory as script

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(senderEmail, password)
        server.sendmail(senderEmail, receiverEmail, text)
    
if __name__ == "__main__":
    # Setting debug to True enables debug output. This line should be
    # removed before deploying a production app.
    application.debug = True
    application.run(host="0.0.0.0", port=80)
    
